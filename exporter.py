import pandas as pd
import numpy as np
import io
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


class BankStatementExporter:
    """
    Export des données bancaires vers Excel ou CSV.
    """

    # Palette de couleurs
    C_HEADER_BG    = "1F4E79"
    C_HEADER_FG    = "FFFFFF"
    C_SOLDE_BG     = "D6EAF8"
    C_SOLDE_FG     = "1A5276"
    C_DEBIT_BG     = "FADBD8"
    C_CREDIT_BG    = "D5F5E3"
    C_ALT_ROW      = "F8F9FA"
    C_BORDER       = "BDBDBD"
    C_TITLE        = "1F4E79"
    C_SUBTITLE     = "2E75B6"

    COLUMN_WIDTHS = {
        'Date':        16,
        'Référence':   14,
        'Libellé':     48,
        'Date_Valeur': 16,
        'Débit':       18,
        'Crédit':      18,
        'Solde':       20,
    }

    def to_excel(self, df: pd.DataFrame,
                 stats: dict = None,
                 account_info: dict = None) -> bytes:
        """
        Génère un fichier Excel formaté.

        Returns:
            bytes du fichier .xlsx
        """
        wb = Workbook()

        # Feuille principale
        ws_data = wb.active
        ws_data.title = "Relevé"
        self._build_data_sheet(ws_data, df, account_info)

        # Feuille résumé
        if stats:
            ws_stats = wb.create_sheet("Résumé")
            self._build_summary_sheet(ws_stats, stats, account_info)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def to_csv(self, df: pd.DataFrame) -> bytes:
        """Génère un fichier CSV UTF-8 avec BOM (compatible Excel)."""
        buf = io.StringIO()
        df.to_csv(buf, index=False, sep=';', encoding='utf-8-sig')
        return buf.getvalue().encode('utf-8-sig')

    # ------------------------------------------------------------------
    # FEUILLE DONNÉES
    # ------------------------------------------------------------------

    def _build_data_sheet(self, ws, df: pd.DataFrame,
                          account_info: dict = None):
        row_ptr = 1

        # Titre
        if account_info:
            row_ptr = self._write_title(ws, account_info, row_ptr)

        # En-têtes colonnes
        headers = list(self.COLUMN_WIDTHS.keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_ptr, column=col_idx, value=header)
            cell.font = Font(
                bold=True, color=self.C_HEADER_FG,
                name='Calibri', size=11
            )
            cell.fill = PatternFill("solid", fgColor=self.C_HEADER_BG)
            cell.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True
            )
            cell.border = self._border()

        ws.row_dimensions[row_ptr].height = 28

        # Filtre automatique
        last_col = get_column_letter(len(headers))
        header_row = row_ptr
        row_ptr += 1

        # Données
        for _, row in df.iterrows():
            is_solde = bool(re.search(
                r'solde', str(row.get('Libellé', '')), re.IGNORECASE
            ) if hasattr(row, 'get') else False)

            values = [
                row.get('Date', ''),
                row.get('Référence', ''),
                row.get('Libellé', ''),
                row.get('Date_Valeur', ''),
                self._num(row.get('Débit')),
                self._num(row.get('Crédit')),
                self._num(row.get('Solde')),
            ]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_ptr, column=col_idx, value=val)
                cell.font = Font(
                    name='Calibri', size=10,
                    bold=is_solde,
                    color=self.C_SOLDE_FG if is_solde else "000000"
                )
                cell.border = self._border('thin')

                # Alignement
                if col_idx in (5, 6, 7):
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = '#,##0'
                elif col_idx == 3:
                    cell.alignment = Alignment(
                        horizontal='left', wrap_text=True
                    )
                else:
                    cell.alignment = Alignment(horizontal='center')

                # Couleur fond
                if is_solde:
                    cell.fill = PatternFill("solid", fgColor=self.C_SOLDE_BG)
                elif col_idx == 5 and val:
                    cell.fill = PatternFill("solid", fgColor=self.C_DEBIT_BG)
                elif col_idx == 6 and val:
                    cell.fill = PatternFill("solid", fgColor=self.C_CREDIT_BG)
                elif row_ptr % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=self.C_ALT_ROW)

            row_ptr += 1

        # Largeurs colonnes
        for col_idx, (col_name, width) in enumerate(
            self.COLUMN_WIDTHS.items(), 1
        ):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Filtre + figer
        last_row = row_ptr - 1
        ws.auto_filter.ref = (
            f"A{header_row}:{last_col}{last_row}"
        )
        ws.freeze_panes = f"A{header_row + 1}"

    def _write_title(self, ws, info: dict, row_ptr: int) -> int:
        """Écrit le bloc titre/info compte."""
        # Titre principal
        ws.merge_cells(f'A{row_ptr}:G{row_ptr}')
        c = ws[f'A{row_ptr}']
        c.value = "FINANCIAL HOUSE S.A — Historique Compte Client"
        c.font = Font(bold=True, size=14, color=self.C_TITLE, name='Calibri')
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row_ptr].height = 30
        row_ptr += 1

        fields = [
            ('Titulaire',  info.get('account_name', '')),
            ('N° Compte',  info.get('account_id', '')),
            ('Période',    info.get('period', '')),
            ('Extraction', info.get('extraction_date', '')),
        ]

        for label, value in fields:
            ws.merge_cells(f'A{row_ptr}:B{row_ptr}')
            ws.merge_cells(f'C{row_ptr}:G{row_ptr}')
            lbl = ws[f'A{row_ptr}']
            lbl.value = label
            lbl.font = Font(bold=True, name='Calibri', size=10)
            lbl.alignment = Alignment(horizontal='right')

            val = ws[f'C{row_ptr}']
            val.value = value
            val.font = Font(name='Calibri', size=10)

            row_ptr += 1

        row_ptr += 1  # ligne vide
        return row_ptr

    # ------------------------------------------------------------------
    # FEUILLE RÉSUMÉ
    # ------------------------------------------------------------------

    def _build_summary_sheet(self, ws, stats: dict,
                              account_info: dict = None):
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 22

        row = 1
        ws.merge_cells(f'A{row}:B{row}')
        c = ws[f'A{row}']
        c.value = "📊 RÉSUMÉ DU COMPTE"
        c.font = Font(bold=True, size=14, color=self.C_TITLE)
        c.alignment = Alignment(horizontal='center')
        ws.row_dimensions[row].height = 28
        row += 2

        sections = [
            ("💰 FLUX FINANCIERS", [
                ("Total Crédits",    stats.get('total_credit', 0)),
                ("Total Débits",     stats.get('total_debit', 0)),
                ("Flux Net",         stats.get('net', 0)),
            ]),
            ("🏦 SOLDES", [
                ("Solde d'ouverture", stats.get('solde_ouverture')),
                ("Solde de clôture",  stats.get('solde_cloture')),
            ]),
            ("📅 PÉRIODE", [
                ("Date de début", stats.get('periode_debut', '')),
                ("Date de fin",   stats.get('periode_fin', '')),
            ]),
            ("📋 TRANSACTIONS", [
                ("Nombre total", stats.get('total_transactions', 0)),
            ]),
        ]

        for section_title, items in sections:
            # Titre de section
            ws.merge_cells(f'A{row}:B{row}')
            c = ws[f'A{row}']
            c.value = section_title
            c.font = Font(bold=True, size=11, color=self.C_HEADER_FG)
            c.fill = PatternFill("solid", fgColor=self.C_SUBTITLE)
            c.alignment = Alignment(horizontal='left',
                                    indent=1, vertical='center')
            ws.row_dimensions[row].height = 22
            row += 1

            for label, value in items:
                lbl = ws.cell(row=row, column=1, value=label)
                lbl.font = Font(name='Calibri', size=10, bold=True)
                lbl.fill = PatternFill("solid", fgColor=self.C_ALT_ROW)
                lbl.border = self._border('thin')

                val = ws.cell(row=row, column=2, value=value)
                val.font = Font(name='Calibri', size=10)
                val.border = self._border('thin')
                val.alignment = Alignment(horizontal='right')
                if isinstance(value, (int, float)):
                    val.number_format = '#,##0'

                row += 1

            row += 1  # espacement

    # ------------------------------------------------------------------
    # UTILITAIRES
    # ------------------------------------------------------------------

    def _border(self, style='medium') -> Border:
        s = Side(style=style, color=self.C_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    def _num(self, val):
        """Convertit en float ou retourne None."""
        if val is None or (isinstance(val, float) and np.isnan(val)):
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None


# Import manquant dans exporter.py
import re
