"""
SKAB Bank Statement Extractor - Edition Comptabilité Odoo 18
Génère CSV + Excel (avec Références, Soldes corrigés, et design de cellules propre)
"""

import streamlit as st
import pandas as pd
import io
import os
import json
import plotly.express as px
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from streamlit_local_storage import LocalStorage

# Modules personnalisés
from extractor_openrouter import OpenRouterExtractor
from cleaner import DataCleaner
from bank_configs import get_bank_list, get_bank_config


# ====================== CONFIGURATION ======================
st.set_page_config(page_title="SKAB Extractor - Export Odoo", page_icon="🏦", layout="wide")

st.markdown("""
<style>
    .main-header { background: linear-gradient(135deg, #1B3A5C, #2E75B6); padding: 2rem; border-radius: 16px; color: white; margin-bottom: 2rem; text-align: center; }
    .main-header h1 { margin: 0; font-size: 1.8rem; }
    .main-header p { margin: 0.5rem 0 0 0; opacity: 0.9; }
    .metric-card {
        background: #ffffff;
        padding: 1rem 1.2rem;
        border-radius: 12px;
        border: 1px solid #E0E0E0;
        box-shadow: 0 1px 4px rgba(0,0,0,0.06);
        text-align: center;
    }
    .metric-card .label { font-size: 0.8rem; color: #666; margin-bottom: 4px; text-transform: uppercase; letter-spacing: 0.5px; }
    .metric-card .value { font-size: 1.6rem; font-weight: 700; color: #1B3A5C; }
    .metric-card .value.credit { color: #2ECC71; }
    .metric-card .value.debit { color: #E74C3C; }
    .metric-card .value.balance { color: #1B3A5C; }
    .metric-card .value.positive { color: #2ECC71; }
    .metric-card .value.negative { color: #E74C3C; }
    .stDataFrame { border-radius: 8px; border: 1px solid #E0E0E0; }
</style>
""", unsafe_allow_html=True)


def get_openrouter_key():
    return st.secrets.get("OPENROUTER_API_KEY", "")


# ====================== PERSISTANCE NAVIGATEUR (localStorage) ======================
LOCAL_STORAGE_KEY = "skab_session_data"
localS = LocalStorage()

def _save_session_to_browser():
    try:
        payload = {
            "df_clean": st.session_state.df_clean.to_json(orient="split", date_format="iso"),
            "stats": st.session_state.stats,
            "banque_selectionnee": st.session_state.banque_selectionnee,
            "uploaded_file_name": st.session_state.get("uploaded_file_name"),
            "failed_pages": st.session_state.failed_pages,
        }
        localS.setItem(LOCAL_STORAGE_KEY, json.dumps(payload), key="skab_set_session")
    except Exception:
        pass

def _clear_session_in_browser():
    try:
        localS.deleteItem(LOCAL_STORAGE_KEY, key="skab_delete_session")
    except Exception:
        pass

def _restore_session_from_browser() -> bool:
    try:
        raw = localS.getItem(LOCAL_STORAGE_KEY, key="skab_get_session")
        if not raw:
            return False
        payload = json.loads(raw) if isinstance(raw, str) else raw
        df_restored = pd.read_json(io.StringIO(payload["df_clean"]), orient="split")
        st.session_state.df_clean = df_restored
        st.session_state.stats = payload.get("stats")
        st.session_state.banque_selectionnee = payload.get("banque_selectionnee", "UNICS")
        st.session_state.uploaded_file_name = payload.get("uploaded_file_name")
        st.session_state.failed_pages = payload.get("failed_pages", [])
        st.session_state.extraction_done = True
        return True
    except Exception:
        return False


if "extraction_done" not in st.session_state:
    st.session_state.update({
        "extraction_done": False,
        "show_confirm": False,
        "df_clean": None,
        "stats": None,
        "banque_selectionnee": "UNICS",
        "pdf_bytes_cache": None,
        "extraction_in_progress": False,
        "total_pages": None,
        "current_page": 1,
        "collected_transactions": [],
        "extraction_method": "vision",
        "failed_pages": [],
        "retry_failed_pages": False,
        "last_known_balance": None,
        "uploaded_file_name": None,
        "browser_restore_done": False,
        "browser_restore_attempts": 0,
    })

if (
    not st.session_state.extraction_done
    and not st.session_state.extraction_in_progress
    and not st.session_state.browser_restore_done
):
    if _restore_session_from_browser():
        st.session_state.browser_restore_done = True
        st.rerun()
    else:
        st.session_state.browser_restore_attempts += 1
        if st.session_state.browser_restore_attempts >= 5:
            st.session_state.browser_restore_done = True

BATCH_SIZE = 8

# ====================== SIDEBAR ======================
with st.sidebar:
    st.title("🏦 SKAB Extractor")
    st.caption("Extraction et conversion de relevés bancaires camerounais")

    uploaded_file = st.file_uploader("📄 Charger le relevé PDF", type=["pdf"])

    BANK_LIST = get_bank_list()
    st.session_state.setdefault("banque_widget", "UNICS")

    if uploaded_file is not None:
        st.session_state.uploaded_file_name = uploaded_file.name
        file_sig = f"{uploaded_file.name}:{uploaded_file.size}"
        if st.session_state.get("bank_detect_sig") != file_sig:
            st.session_state.bank_detect_sig = file_sig
            if get_openrouter_key():
                with st.spinner("🔍 Détection automatique de la banque..."):
                    try:
                        detector = OpenRouterExtractor(
                            api_key=get_openrouter_key(), mode="vision",
                            banque_nom="Autre banque", verbose_debug=False,
                        )
                        detected = detector.detect_bank(uploaded_file.getvalue())
                    except Exception:
                        detected = None
                if detected:
                    st.session_state.banque_widget = detected
                    st.success(f"🏦 Banque détectée : **{detected}**")
                else:
                    st.info("Banque non détectée automatiquement — sélectionnez-la ci-dessous.")

    banque_sel = st.selectbox("🏦 Banque émettrice", BANK_LIST, key="banque_widget")
    st.session_state.banque_selectionnee = banque_sel

    method = st.radio("🔍 Méthode d'analyse", ["vision", "hybrid"],
                      help="Vision : analyse d'images | Hybride : texte puis vision si nécessaire")

    if not get_openrouter_key():
        st.warning("""
            ⚠️ **Clé API manquante**  
            Ajoutez `OPENROUTER_API_KEY` dans  
            `.streamlit/secrets.toml`
        """)

    if st.button("🔄 Nouvelle extraction", use_container_width=True):
        _clear_session_in_browser()
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.rerun()

# ====================== HEADER ======================
st.markdown('<div class="main-header"><h1>🏦 SKAB Bank Statement Extractor</h1><p>Génération de fichiers d\'importation pour la comptabilité Odoo</p></div>', unsafe_allow_html=True)

if (
    not st.session_state.extraction_done
    and not st.session_state.extraction_in_progress
    and not uploaded_file
):
    col_restore, _ = st.columns([1, 3])
    with col_restore:
        if st.button("🔁 Restaurer la dernière session", use_container_width=True):
            st.session_state.browser_restore_done = False
            st.session_state.browser_restore_attempts = 0
            st.rerun()

# ====================== EXTRACTION ======================
if (
    uploaded_file
    and not st.session_state.extraction_done
    and not st.session_state.show_confirm
    and not st.session_state.extraction_in_progress
):
    if st.button("🔍 Analyser le relevé", type="primary", use_container_width=True):
        st.session_state.pdf_bytes_cache = uploaded_file.read()
        st.session_state.show_confirm = True
        st.rerun()

if st.session_state.show_confirm:
    if not get_openrouter_key():
        st.error("❌ Clé API OpenRouter manquante. Configurez-la dans les secrets Streamlit.")
    else:
        col1, col2 = st.columns([3, 1])
        with col1:
            st.info(f"**Analyse imminente** — Banque: **{st.session_state.banque_selectionnee}**")
        with col2:
            if st.button("✅ Confirmer l'analyse", type="primary", use_container_width=True):
                st.session_state.extraction_method = method
                st.session_state.extraction_in_progress = True
                st.session_state.browser_restore_done = True
                st.session_state.show_confirm = False
                st.session_state.total_pages = None
                st.session_state.current_page = 1
                st.session_state.collected_transactions = []
                st.session_state.failed_pages = []
                st.session_state.last_known_balance = None
                st.rerun()

# ====================== EXTRACTION PAR LOTS ======================
if st.session_state.extraction_in_progress:
    if not get_openrouter_key():
        st.error("❌ Clé API OpenRouter manquante.")
        st.session_state.extraction_in_progress = False
    else:
        try:
            extractor = OpenRouterExtractor(
                api_key=get_openrouter_key(),
                mode=st.session_state.extraction_method,
                banque_nom=st.session_state.banque_selectionnee,
                verbose_debug=False,
            )

            if st.session_state.extraction_method != "vision":
                with st.spinner("Extraction en cours (mode hybride)..."):
                    df_raw = extractor.extract(st.session_state.pdf_bytes_cache)
                st.session_state.failed_pages = sorted(set(extractor.failed_pages))
                cleaner = DataCleaner()
                df_clean = cleaner.clean(df_raw, banque_nom=st.session_state.banque_selectionnee)
                df_clean = cleaner.check_consistency(df_clean)
                st.session_state.df_clean = df_clean
                st.session_state.stats = cleaner.get_statistics(df_clean, banque_nom=st.session_state.banque_selectionnee)
                st.session_state.extraction_done = True
                st.session_state.extraction_in_progress = False
                _save_session_to_browser()
                st.rerun()
            else:
                if st.session_state.total_pages is None:
                    st.session_state.total_pages = extractor.get_total_pages(st.session_state.pdf_bytes_cache)
                    if not st.session_state.total_pages:
                        st.error("❌ Impossible de lire le PDF (nombre de pages introuvable).")
                        st.session_state.extraction_in_progress = False
                        st.stop()
                    st.rerun()

                total_pages = st.session_state.total_pages
                current = st.session_state.current_page
                batch_end = min(current + BATCH_SIZE - 1, total_pages)

                st.progress(
                    (current - 1) / total_pages,
                    text=f"🔍 Analyse en cours — pages {current} à {batch_end} sur {total_pages}...",
                )

                batch_transactions = extractor.extract_transactions(
                    st.session_state.pdf_bytes_cache, current, batch_end, total_pages,
                    starting_balance=st.session_state.last_known_balance,
                )
                st.session_state.collected_transactions.extend(batch_transactions)
                st.session_state.last_known_balance = extractor.last_balance_hint
                st.session_state.failed_pages = sorted(
                    set(st.session_state.failed_pages) | set(extractor.failed_pages)
                )
                st.session_state.current_page = batch_end + 1

                if st.session_state.current_page > total_pages:
                    df_raw = extractor.build_dataframe(st.session_state.collected_transactions)
                    cleaner = DataCleaner()
                    df_clean = cleaner.clean(df_raw, banque_nom=st.session_state.banque_selectionnee)
                    df_clean = cleaner.check_consistency(df_clean)
                    st.session_state.df_clean = df_clean
                    st.session_state.stats = cleaner.get_statistics(df_clean, banque_nom=st.session_state.banque_selectionnee)
                    st.session_state.extraction_done = True
                    st.session_state.extraction_in_progress = False
                    _save_session_to_browser()

                st.rerun()

        except Exception as e:
            st.error(f"❌ Erreur lors de l'extraction : {str(e)}")
            st.session_state.extraction_in_progress = False

# ====================== NOUVEL ESSAI ======================
if st.session_state.retry_failed_pages:
    if not get_openrouter_key():
        st.error("❌ Clé API OpenRouter manquante.")
        st.session_state.retry_failed_pages = False
    else:
        try:
            extractor = OpenRouterExtractor(
                api_key=get_openrouter_key(),
                mode="vision",
                banque_nom=st.session_state.banque_selectionnee,
                verbose_debug=False,
            )
            pages_a_reessayer = list(st.session_state.failed_pages)
            hint = st.session_state.collected_transactions[-1].get("solde") if st.session_state.collected_transactions else None
            with st.spinner(f"Nouvel essai sur {len(pages_a_reessayer)} page(s)..."):
                new_transactions = extractor.extract_specific_pages(
                    st.session_state.pdf_bytes_cache, pages_a_reessayer, st.session_state.total_pages,
                    starting_balance=hint,
                )
            st.session_state.collected_transactions.extend(new_transactions)
            st.session_state.failed_pages = sorted(set(extractor.failed_pages))

            df_raw = extractor.build_dataframe(st.session_state.collected_transactions)
            cleaner = DataCleaner()
            df_clean = cleaner.clean(df_raw, banque_nom=st.session_state.banque_selectionnee)
            df_clean = cleaner.check_consistency(df_clean)
            st.session_state.df_clean = df_clean
            st.session_state.stats = cleaner.get_statistics(df_clean, banque_nom=st.session_state.banque_selectionnee)
            st.session_state.retry_failed_pages = False
            _save_session_to_browser()
            st.rerun()
        except Exception as e:
            st.error(f"❌ Erreur lors du nouvel essai : {str(e)}")
            st.session_state.retry_failed_pages = False

# ====================== RÉSULTATS ======================
if st.session_state.extraction_done and st.session_state.df_clean is not None:
    df_display = st.session_state.df_clean.copy()

    df_display['Date'] = pd.to_datetime(df_display['Date'], dayfirst=True, errors='coerce')
    df_display = df_display.dropna(subset=['Date'])

    if st.session_state.failed_pages:
        pages_str = ", ".join(str(p) for p in st.session_state.failed_pages)
        col_warn, col_btn = st.columns([4, 1])
        with col_warn:
            st.error(f"⚠️ {len(st.session_state.failed_pages)} page(s) n'ont pas pu être lues et ne sont PAS incluses : page(s) {pages_str}.")
        with col_btn:
            if st.session_state.extraction_method == "vision" and st.button("🔄 Relancer ces pages", use_container_width=True):
                st.session_state.retry_failed_pages = True
                st.rerun()

    stats = st.session_state.stats or {}

    col1, col2, col3 = st.columns(3)
    with col1:
        so = stats.get('solde_ouverture')
        so_class = "positive" if (so is not None and so >= 0) else "negative"
        so_val = f"{so:,.0f} FCFA" if so is not None else "N/A"
        st.markdown(f"""
        <div class="metric-card">
            <div class="label">Solde d'ouverture (relevé)</div>
            <div class="value {so_class}">{so_val}</div>
        </div>""", unsafe_allow_html=True)
    
    with col2:
        sc = stats.get('solde_cloture')
        sc_class = "positive" if (sc is not None and sc >= 0) else "negative"
        sc_val = f"{sc:,.0f} FCFA" if sc is not None else "N/A"
        st.markdown(f"""
        <div class="metric-card">
            <div class="label">Solde de clôture (relevé)</div>
            <div class="value {sc_class}">{sc_val}</div>
        </div>""", unsafe_allow_html=True)
    
    with col3:
        net = stats.get('net', 0)
        net_class = "positive" if net >= 0 else "negative"
        st.markdown(f"""
        <div class="metric-card">
            <div class="label">Flux net</div>
            <div class="value {net_class}">{net:,.0f} FCFA</div>
        </div>""", unsafe_allow_html=True)

    st.caption("🔎 La logique d'Overlap garantit qu'aucune transaction n'est coupée et que les faux calculs d'IA sont bloqués.")

    ecart_oc = stats.get('ecart_ouverture_cloture')
    if ecart_oc is None:
        if stats.get('solde_ouverture') is None or stats.get('solde_cloture') is None:
            st.info("ℹ️ Contrôle de cohérence indisponible : soldes d'ouverture/clôture non détectés.")
    elif abs(ecart_oc) <= 1:
        st.success("✅ Cohérence parfaite vérifiée : solde d'ouverture + flux net = solde de clôture attendu.")
    else:
        st.warning(f"⚠️ Écart mathématique détecté de {ecart_oc:,.0f} FCFA entre le solde de clôture imprimé et les mouvements extraits.")
    
    col4, col5, col6 = st.columns(3)
    with col4:
        st.markdown(f'<div class="metric-card"><div class="label">Total crédits</div><div class="value credit">{stats.get("total_credit", 0):,.0f} FCFA</div></div>', unsafe_allow_html=True)
    with col5:
        st.markdown(f'<div class="metric-card"><div class="label">Total débits</div><div class="value debit">{stats.get("total_debit", 0):,.0f} FCFA</div></div>', unsafe_allow_html=True)
    with col6:
        st.markdown(f'<div class="metric-card"><div class="label">Lignes extraites</div><div class="value balance">{len(df_display)}</div></div>', unsafe_allow_html=True)

    st.subheader("📊 Flux de trésorerie")
    df_display['Débit'] = pd.to_numeric(df_display['Débit'], errors='coerce').fillna(0)
    df_display['Crédit'] = pd.to_numeric(df_display['Crédit'], errors='coerce').fillna(0)
    df_display['Solde_cumulé'] = df_display['Crédit'].cumsum() - df_display['Débit'].cumsum()

    df_chart = df_display.groupby('Date').agg({'Débit': 'sum', 'Crédit': 'sum', 'Solde_cumulé': 'last'}).reset_index()

    fig = px.bar(df_chart, x='Date', y=['Crédit', 'Débit'], title="Mouvements bancaires", barmode='group', color_discrete_map={"Crédit": "#2ECC71", "Débit": "#E74C3C"}, height=400)
    fig.add_scatter(x=df_chart['Date'], y=df_chart['Solde_cumulé'], mode='lines+markers', name='Solde', line=dict(color="#1B3A5C", width=3), marker=dict(size=6), yaxis='y')
    fig.update_layout(hovermode="x unified", yaxis_title="Montant (FCFA)", legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1), margin=dict(l=40, r=20, t=40, b=40), font=dict(size=12))
    st.plotly_chart(fig, use_container_width=True)

    st.divider()
    st.subheader("📋 Données extraites")

    date_min = df_display['Date'].min().date()
    date_max = df_display['Date'].max().date()
    col_d1, col_d2 = st.columns(2)
    with col_d1:
        date_debut = st.date_input("Du", value=date_min, min_value=date_min, max_value=date_max)
    with col_d2:
        date_fin = st.date_input("Au", value=date_max, min_value=date_min, max_value=date_max)

    if date_debut > date_fin:
        st.warning("⚠️ La date de début est postérieure à la date de fin.")
        df_filtered = df_display.iloc[0:0].copy()
    else:
        df_filtered = df_display[(df_display['Date'].dt.date >= date_debut) & (df_display['Date'].dt.date <= date_fin)].copy()

    # Formater les dates pour Streamlit sans affecter l'objet pandas réel
    df_visual = df_filtered.copy()
    df_visual['Date'] = df_visual['Date'].dt.strftime('%d/%m/%Y')
    
    display_cols = ['Date', 'Référence', 'Libellé', 'Débit', 'Crédit', 'Solde', 'Écart']
    display_df = df_visual[[c for c in display_cols if c in df_visual.columns]]
    st.dataframe(display_df, use_container_width=True, height=400)

    # --- EXPORT ---
    st.divider()
    st.subheader("💾 Export")
    
    col_csv, col_xlsx = st.columns(2)

    with col_csv:
        odoo_export = df_filtered.copy()
        odoo_export['Date'] = odoo_export['Date'].dt.strftime('%Y-%m-%d')
        odoo_export = odoo_export.rename(columns={'Date': 'date', 'Libellé': 'payment_ref', 'Référence': 'ref'})
        odoo_export['amount'] = odoo_export['Crédit'].fillna(0) - odoo_export['Débit'].fillna(0)
        odoo_export['ref'] = odoo_export['ref'].replace(0, '').replace('0.0', '')

        csv_cols = ['date', 'payment_ref', 'amount', 'ref', 'Solde'] if 'Solde' in odoo_export.columns else ['date', 'payment_ref', 'amount', 'ref']
        final_csv = odoo_export[[c for c in csv_cols if c in odoo_export.columns]]
        csv_buffer = io.StringIO()
        final_csv.to_csv(csv_buffer, index=False, encoding='utf-8-sig', sep=',')
        
        st.download_button(label="📥 Télécharger CSV", data=csv_buffer.getvalue(), file_name=f"EXPORT_{st.session_state.banque_selectionnee}.csv", mime="text/csv", type="primary", use_container_width=True)

    with col_xlsx:
        excel_buffer = io.BytesIO()
        sheet1_cols = ['Date', 'Référence', 'Libellé', 'Débit', 'Crédit', 'Solde']
        sheet1_df = df_filtered[[c for c in sheet1_cols if c in df_filtered.columns]].reset_index(drop=True)

        bank_cfg = get_bank_config(st.session_state.banque_selectionnee)
        solde_pattern = "|".join(bank_cfg.solde_ouverture_patterns + bank_cfg.solde_cloture_patterns) or (r"ouverture|opening|cl[ôo]ture|cloture|solde\s+d[ée]but|solde\s+de\s+d[ée]but|solde\s+au\s+\d{1,2}|report\s+solde")
        
        if 'Libellé' in sheet1_df.columns:
            is_solde_row = sheet1_df['Libellé'].astype(str).str.lower().str.contains(solde_pattern, na=False, regex=True)
        else:
            is_solde_row = pd.Series(False, index=sheet1_df.index)

        solde_series = pd.to_numeric(sheet1_df.get('Solde'), errors='coerce') if 'Solde' in sheet1_df.columns else pd.Series(dtype=float)
        montant_series = pd.to_numeric(sheet1_df.get('Crédit'), errors='coerce').fillna(0) - pd.to_numeric(sheet1_df.get('Débit'), errors='coerce').fillna(0)
        montant_series = montant_series.where(~is_solde_row, other=pd.NA)

        opening_balance = 0.0
        if len(sheet1_df) and len(solde_series) and pd.notna(solde_series.iloc[0]):
            if is_solde_row.iloc[0]:
                opening_balance = float(solde_series.iloc[0])
            else:
                first_montant = 0.0 if pd.isna(montant_series.iloc[0]) else float(montant_series.iloc[0])
                opening_balance = float(solde_series.iloc[0]) - first_montant

        wb = Workbook()
        ws = wb.active
        ws.title = "RELEVE"

        # Design Excel structuré
        headers = ["Date", "Référence", "Libellé", "Montant", "Solde courant"]
        header_fill = PatternFill("solid", fgColor="1F3864")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'), 
            right=Side(style='thin', color='D3D3D3'), 
            top=Side(style='thin', color='D3D3D3'), 
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align
            c.border = thin_border

        row_ptr = 2
        for i in range(len(sheet1_df)):
            date_val = sheet1_df.iloc[i].get('Date')
            ref_val = sheet1_df.iloc[i].get('Référence', '')
            libelle = sheet1_df.iloc[i].get('Libellé', '')
            is_solde = bool(is_solde_row.iloc[i])
            montant_val = montant_series.iloc[i]

            c_date = ws.cell(row=row_ptr, column=1, value=date_val)
            c_date.number_format = 'dd/mm/yyyy'
            c_date.border = thin_border

            c_ref = ws.cell(row=row_ptr, column=2, value=ref_val)
            c_ref.border = thin_border
            
            c_lib = ws.cell(row=row_ptr, column=3, value=libelle)
            c_lib.border = thin_border

            c_montant = ws.cell(row=row_ptr, column=4)
            c_montant.border = thin_border
            if not is_solde and pd.notna(montant_val):
                c_montant.value = float(montant_val)
                c_montant.number_format = '#,##0;-#,##0'

            c_solde = ws.cell(row=row_ptr, column=5)
            c_solde.border = thin_border
            if is_solde and pd.notna(solde_series.iloc[i]):
                c_solde.value = float(solde_series.iloc[i])
            elif row_ptr == 2:
                c_solde.value = f"={opening_balance:.0f}+D{row_ptr}"
            else:
                c_solde.value = f"=E{row_ptr - 1}+D{row_ptr}"
                
            c_solde.number_format = '#,##0;-#,##0'
            row_ptr += 1

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 65
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 18
        ws.freeze_panes = "A2"

        wb.save(excel_buffer)
        excel_buffer.seek(0)

        source_pdf_name = st.session_state.get("uploaded_file_name") or (uploaded_file.name if uploaded_file else "EXPORT.pdf")
        excel_file_name = f"{os.path.splitext(source_pdf_name)[0]}.xlsx"

        st.download_button(label="📥 Télécharger Excel", data=excel_buffer.getvalue(), file_name=excel_file_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary", use_container_width=True)

    st.success("✅ Fichiers finaux générés et structurés avec succès !")
